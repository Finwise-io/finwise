#!/usr/bin/env python3
"""Adds the retirement-odds METHOD & ASSUMPTIONS tab to the master PRD (founder ask 2026-08-11).

Every number the Monte Carlo runs on, where it comes from, whether the user can change it, and the
honest limits — so the model can be reviewed without reading the code, and so a future change to a
figure has a place it must be recorded.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

PRD = 'docs/FCC-core-55-70/FCC-core-PRD-v1.1-2026-07-01.xlsx'
HEAD = PatternFill('solid', fgColor='085041')
SUB = PatternFill('solid', fgColor='DFF2E9')
WARN = PatternFill('solid', fgColor='FAEEDA')
HF = Font(color='FFFFFF', bold=True, size=11)

# (Assumption, Value, Where it comes from, User can change?, Code)
METHOD = [
    ('METHOD', '', '', '', ''),
    ('Simulation type', 'Monte Carlo — 500 independent paths',
     'Each path plays the whole plan year by year: growth while working, then drawdown to the horizon. A path SUCCEEDS if the balance never reaches zero.',
     'No', 'simulate() · domain/retirement/index.ts'),
    ('The probability', 'successes ÷ 500, rounded to a whole percent',
     'If the money lasts in 420 of 500 paths, the odds read 84%.', 'No', 'chance_of_success'),
    ('Yearly return draw', 'Normal distribution (Box–Muller) around the mean return, spread by the volatility',
     'Each simulated year draws its own return. This is what makes the 500 paths differ.', 'No', 'normal(rng, mean, sd)'),
    ('Randomness', 'Seeded (seed 12345) — deterministic',
     'The same inputs always give the same odds, so the number never jitters between visits. Only changing an assumption changes it.',
     'No', 'mulberry32(seed)'),
    ('Verdict words', '≥80% Likely · 60–79% Uncertain · <60% Unlikely (retired: Holding · Watch closely · Running short)',
     'The word pairs the number so colour is never the only signal.', 'No', 'lensChanceWord / chanceWord'),

    ('GROWTH — annual, nominal, before inflation', '', 'Weighted across YOUR holdings by the amount earmarked for retirement — never one flat rate.', '', 'blendedReturn(accounts)'),
    ('Shares / ETFs', '10.4%', 'S&P 500 total return, 30-year', 'Yes — Plan → Retirement', 'ASSET_KINDS.ret'),
    ('401(k) / Traditional IRA / Roth IRA / HSA', '7.9%', 'Assumed 60/40 mix, 30-year (ESTIMATE)', 'Yes', 'ASSET_KINDS.ret'),
    ('Brokerage (mix unknown)', '8.0%', 'Blended taxable portfolio (ESTIMATE)', 'Yes', 'ASSET_KINDS.ret'),
    ('Bonds & CDs / fixed income', '4.2%', 'US Aggregate Bond index, 30-year', 'Yes', 'ASSET_KINDS.ret'),
    ('High-yield savings · money market · CD · cash management', '4.2% · 4.5% · 4.2% · 3.5%', 'Current rates (ESTIMATE)', 'Yes', 'ASSET_KINDS.ret'),
    ('Savings · checking', '2.4% · 0.5%', '3-month Treasury bill, 30-year · interest checking (ESTIMATE)', 'Yes', 'ASSET_KINDS.ret'),
    ('Gold / commodities', '8.2%', 'Gold (SPDR GLD), 30-year', 'Yes', 'ASSET_KINDS.ret'),
    ('Private equity · hedge funds', '13% · 6%', 'Cambridge Associates US PE ~25-yr · HFRI Fund Weighted ~10-yr (both ESTIMATES)', 'Yes', 'ASSET_KINDS.ret'),
    ('Crypto', '8%', 'NO long-run benchmark exists — a placeholder the user should set (ESTIMATE)', 'Yes', 'ASSET_KINDS.ret'),
    ('Home · vehicle', '4.5% · −5%', 'Case-Shiller US home price, 30-yr · vehicle depreciation (ESTIMATE). Property is excluded from the nest egg.', 'Yes', 'ASSET_KINDS.ret'),
    ('Nothing held yet', '6.0%', 'Fallback so a plan can still be simulated before any account exists', 'Yes', 'blendedReturn fallback'),

    ('THE UPS AND DOWNS — annual standard deviation', '', 'ADDED 2026-08-11. Was previously derived as return × 1.7 (floored 5%) — a rule of thumb that called a savings account 5% volatile and could not tell bonds from shares.', '', 'blendedVolatility(accounts)'),
    ('Shares / ETFs', '±18%', 'S&P 500 annual std dev, 30-year', 'No (derived from holdings)', 'KIND_VOLATILITY'),
    ('401(k) / IRA / Roth / HSA', '±11%', '60/40 mix annual std dev, 30-year (ESTIMATE)', 'No', 'KIND_VOLATILITY'),
    ('Brokerage', '±14%', 'Blended taxable portfolio (ESTIMATE)', 'No', 'KIND_VOLATILITY'),
    ('Bonds & CDs', '±6%', 'US Aggregate Bond index annual std dev', 'No', 'KIND_VOLATILITY'),
    ('Cash (checking · savings · HYSA · MMF · CD)', '±0.5% to ±1.2%', '3-month Treasury bill annual std dev', 'No', 'KIND_VOLATILITY'),
    ('Gold / commodities · private equity · hedge funds', '±16% · ±20% · ±8%', 'Same series as their returns. PE is SMOOTHED — its true swing is wider (ESTIMATE).', 'No', 'KIND_VOLATILITY'),
    ('Crypto', '±70%', 'No long-run series — a deliberately wide placeholder (ESTIMATE)', 'No', 'KIND_VOLATILITY'),
    ('Home', '±10%', 'Case-Shiller annual std dev', 'No', 'KIND_VOLATILITY'),
    ('Nothing held yet', '±12%', 'Fallback; the scenario sliders instead derive volatility from the typed return (return × 1.7, clamped 5–20%)', 'No', 'volOf() — scenario only'),

    ('YOUR ANSWERS', '', '', '', ''),
    ('Inflation', '2.5% a year', 'Default; raises spending AND Social Security over time', 'Yes — Plan → Retirement', 'inflationRate'),
    ('Retirement age', 'Your answer, else 65', 'A retiree is simulated from today', 'Yes', 'targetRetirementAge / assumptions.retireAge'),
    ('Plan-to age (horizon)', 'Your answer, else 90', 'The age the money must reach', 'Yes', 'horizonAge'),
    ('Monthly spending in retirement', 'Your answer, else today\'s spending, else $5,000', 'Inflated to the retirement year, then each year after', 'Yes', 'retirementSpendMonthly'),
    ('Social Security / pension', 'Your captured amount, adjusted for claim age', 'Counted only FROM the claim age — 67 unless chosen otherwise', 'Yes — Plan → Social Security timing', 'ssBenefitAtClaimAge'),
    ('Contributions', 'From your onboarding answers, grown yearly', 'Added each working year', 'Yes', 'monthlyContributionsFromOnboarding'),
    ('Nest egg', 'Retirement-EARMARKED balances only', 'Cash and property default to 0% earmarked; investment and retirement accounts to 100%. Per-account override available.', 'Yes — per account', 'retirementEarmarkedValue'),

    ('TAX & RULES INSIDE THE SIMULATION', '', '', '', ''),
    ('Required withdrawals (RMD)', 'From age 73, IRS divisor by age', 'Forced out of the pre-tax slice; the NET stays in the pot', 'No', 'RMD_START_AGE_SIM / rmdDivisorSim'),
    ('Tax on required withdrawals', '22% of the RMD leaves the pot', 'Default effective rate; the pre-tax share is computed from the actual account mix', 'Yes', 'rmd_tax_rate · pre_tax_share'),
    ('Big one-time costs', 'Your captured costs, in the year given', 'Inflated from today\'s dollars to their year; costs outside the horizon are ignored', 'Yes — Plan → Big costs', 'one_off_costs'),
    ('"Needed" figure (the gap)', '25 × net annual spend', 'The 4% rule, applied to spending net of guaranteed income. Used for the gap and the suggested extra monthly saving, NOT for the odds.', 'No', 'needed = spend × 25'),

    ('HONEST LIMITS — stated in the info dot too', '', '', '', ''),
    ('Independent years', 'Each year\'s return is drawn independently',
     'Real markets cluster — bad years arrive together, and clustering early in retirement is what breaks a plan (sequence-of-returns risk). Independent draws make the odds slightly BETTER than reality for someone about to retire.',
     'No', 'Known limitation'),
    ('Holdings move together', 'Volatility is a weighted average across holdings',
     'This ignores diversification, so it slightly OVERSTATES risk — the odds come out cautious rather than flattering. The opposite side of the previous limit.',
     'No', 'blendedVolatility'),
    ('Not a promise', 'An estimate from your numbers and these assumptions',
     'Every figure above is visible to the user in the retirement info dot, in plain English, with the screen that changes it.',
     '—', 'GLOSSARY.nestEggMath'),
]

wb = openpyxl.load_workbook(PRD)
title = 'Retirement odds — assumptions'
if title in wb.sheetnames:
    del wb[title]
ws = wb.create_sheet(title)
heads = ['Assumption', 'Value', 'Where it comes from / what it means', 'User can change?', 'Code']
ws.append(heads)
for c in range(1, len(heads) + 1):
    ws.cell(1, c).fill = HEAD
    ws.cell(1, c).font = HF
ws.row_dimensions[1].height = 22

for row in METHOD:
    ws.append(list(row))
    r = ws.max_row
    for c in range(1, 6):
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    if row[1] == '' and row[2] in ('', None) or (row[1] == '' and row[0].isupper()):
        for c in range(1, 6):
            ws.cell(r, c).fill = SUB
            ws.cell(r, c).font = Font(bold=True)
    elif row[0].startswith(('Independent years', 'Holdings move together', 'Not a promise')):
        for c in range(1, 6):
            ws.cell(r, c).fill = WARN

for col, w in zip('ABCDE', [44, 30, 92, 26, 34]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'

# point the Amendments tab at it, so the change is dated in the PRD's own log
am = wb['Amendments v1.1→v1.2']
am.append(['2026-08-11', 'F6/F9 Retirement odds — method & assumptions',
           'NEW TAB "Retirement odds — assumptions": every input the Monte Carlo runs on, its value, its source, and whether the user can change it. CHANGED: portfolio volatility is now measured per asset class from the same long-run series as the returns (KIND_VOLATILITY + blendedVolatility), replacing the derived "return × 1.7" rule which called a savings account 5% volatile and could not distinguish bonds from shares. The old rule remains only as the scenario fallback when a return is typed with nothing held. ALSO: the retirement info dot now explains the whole method and every one of these numbers in plain English, and names the screen that changes them.',
           'Founder ask 2026-08-11: "where are we getting data or assumptions for monte carlo simulation?" → fix the weak assumption, explain it to a non-finance reader, and record it in the master PRD.'])
for c in range(1, 5):
    am.cell(am.max_row, c).alignment = Alignment(wrap_text=True, vertical='top')

wb.save(PRD)
print('wrote tab:', title, '·', ws.max_row - 1, 'rows | Amendments now', am.max_row - 1, 'rows')
