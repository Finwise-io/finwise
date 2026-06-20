# Regenerates docs/finwise-tracker.xlsx (consolidated bugs + backlog + shipped).
# Usage:  python3 -m venv .venv && .venv/bin/pip install openpyxl && .venv/bin/python scripts/gen-tracker.py
# Edit the curated backlog (list B) + shipped (list S) here; the Bug Ledger tab auto-parses finwise-bug-ledger.md.

import re, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LEDGER = "/Users/palakjain/Vibe_Coding/finwise/docs/finwise-bug-ledger.md"
OUT = "/Users/palakjain/Vibe_Coding/finwise/docs/finwise-tracker.xlsx"

# ---- styling helpers ----
HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILLS = {
    "blocker": PatternFill("solid", fgColor="F8CBCB"),
    "next":    PatternFill("solid", fgColor="CFE2F3"),
    "later":   PatternFill("solid", fgColor="EAEAEA"),
    "done":    PatternFill("solid", fgColor="D9EAD3"),
    "partial": PatternFill("solid", fgColor="FCE5CD"),
}
def status_fill(s):
    s = (s or "").lower()
    if any(k in s for k in ["blocker", "🚧"]): return FILLS["blocker"]
    if any(k in s for k in ["fixed", "done", "shipped", "resolved", "✅", "🟢"]): return FILLS["done"]
    if any(k in s for k in ["partial", "🟡", "in progress", "in-progress", "~"]): return FILLS["partial"]
    if any(k in s for k in ["next", "🔵"]): return FILLS["next"]
    if any(k in s for k in ["later", "⚪", "deferred", "by-design"]): return FILLS["later"]
    return None

def write_sheet(ws, headers, rows, widths, status_col=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True); cell.border = BORDER
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = WRAP if widths[c-1] >= 30 else TOP
            cell.border = BORDER
            cell.font = Font(size=10)
        if status_col is not None:
            f = status_fill(row[status_col])
            if f: ws.cell(r, status_col + 1).fill = f
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

wb = Workbook()

# ============ SHEET 1: OPEN ITEMS & BACKLOG (curated) ============
ws1 = wb.active; ws1.title = "Open Items & Backlog"
# id, category, item, priority, status, owner, next step
B = [
 # ---- LAUNCH BLOCKERS ----
 ["L-1","Launch","Market-data licensing — replace the unofficial Yahoo endpoint with a licensed end-of-day vendor (Tiingo / EODHD / Twelve Data)","Blocker","Open","You + Eng","Pick a vendor + get an API key; then a ~1-file PriceProvider swap in marketData.ts. THE main remaining launch blocker."],
 ["L-2","Launch","App Store listing — screenshots (6.7\"/6.5\"), description, keywords, category, age rating","Blocker","Open","You","Prepare in App Store Connect."],
 ["L-3","Launch","App Privacy 'nutrition label' must match the privacy manifest (Email, Name, Financial info, Crash data; all app-functionality, no tracking)","Blocker","Open","You","Enter in App Store Connect; Apple cross-checks against the manifest."],
 ["L-4","Launch","Reviewer demo login — a working email+password with sample data in the submission notes","Blocker","Open","You","Without it the reviewer can't get past login → rejection."],
 ["L-5","Launch","On-device test of new flows (delete account, recovery code, Face ID lock) on the TestFlight build","Blocker (verify)","In progress","You","TestFlight build #21 (v1.0.1) uploaded 2026-06-20; test once it finishes processing."],
 # ---- SECURITY / PRIVACY ----
 ["S-1","Security/Privacy","Household / partner sharing under zero-knowledge encryption","Next-up","Open","Eng","Two passwords can't share one key today; fails safe (no crash). Restore via a shared key passed through the invite code (key-wrapping)."],
 ["S-2","Security/Privacy","Multi-factor authentication (MFA)","Later","Open","Eng","App lock + inactivity timeout already shipped; MFA is the remaining auth-hardening item."],
 ["S-3","Security/Privacy","Sentry crash reporting — wire @sentry/react-native properly","Later","Open","Eng","Needs Metro config (getSentryExpoConfig) + Expo plugin + DSN. Global JS error handler already ships crash capture locally."],
 ["S-4","Security/Privacy","Deploy the AI proxy (functions/) + set AI_PROXY_URL — or ship v1 without AI tips","Next-up","Open","You + Eng","Needs the Firebase Blaze plan. App already falls back to on-device tips."],
 ["S-5","Security/Privacy","Dependency audit — triage 58 npm advisories (npm audit --omit=dev)","Later","Open","Eng","Almost all are build-time tooling, not shipped in the app."],
 # ---- ACCESSIBILITY ----
 ["A-1","Accessibility","Per-screen screen-reader labels on the remaining ~26 screens","Next-up","Open","Eng/UI","Shared Button/controls, tab bar, TopBar, Home + new screens already labelled."],
 ["A-2","Accessibility","Dark mode","Later","Open","UI",""],
 ["A-3","Accessibility","Privacy-blur when the app is backgrounded","Later","Open","Eng","Hide financial data in the app switcher."],
 ["A-4","Accessibility","Color-contrast audit (G-32)","Later","Open","UI",""],
 # ---- FEATURES / DATA ----
 ["F-1","Features","Receipt OCR native rebuild + test (ML Kit)","Next-up","Open","Eng","npx expo run:ios to activate, then test a real receipt."],
 ["F-2","Features","Plaid bank / brokerage linking","Later","Open","Eng","Biggest stickiness lever; manual entry doesn't scale. Needs Plaid keys + a backend."],
 ["F-3","Features","Savings goals — finish the waterfall-backed goals screen UI","Next-up","Open","Eng",""],
 ["F-4","Features","Drawdown deeper tax-awareness (Roth conversions, capital-gains harvesting, catch-up)","Later","Open","Eng","Drawdown v1 shipped (withdrawal order, depletion age, RMDs)."],
 ["F-5","Features","Single-total vs categories spend reconciliation — one source of truth","Next-up","Open","Eng",""],
 ["F-6","Features","Full string internationalization (translate copy, not just numbers)","Later","Open","Eng","Currency picker already reformats numbers."],
 ["F-7","Features","Auto-derive dividends from ticker; ticker autocomplete; price-cache TTL","Later","Open","Eng","Perf v2 polish."],
 ["F-8","Features","Life-stage gaps: 529 college planning, life/LTC insurance, estate/legacy","Later","Open","Eng",""],
 ["F-9","Features","Push notifications (reminders/nudges) — plugin configured, needs APNs + logic","Later","Open","Eng",""],
 # ---- POLISH / INFRA ----
 ["P-1","Polish","Net-worth-over-time card on Home; runway / emergency-fund insight","Later","Open","UI",""],
 ["P-2","Polish","Header + bottom tab bar redesign; legacy detail-screen redesigns","Later","Open","UI",""],
 ["I-1","Infra","GitHub Pages 'pages build and deployment' workflow keeps failing","Later","Open","You/Eng","Harmless; disable Pages in repo settings or fix the Jekyll build over docs/."],
]
write_sheet(ws1, ["ID","Category","Item","Priority","Status","Owner","Next step / detail"], B,
            [7,16,46,14,12,12,60], status_col=3)

# ============ SHEET 2: BUG LEDGER (parsed from markdown) ============
ws2 = wb.create_sheet("Bug Ledger (all)")
bug_rows = []
with open(LEDGER) as f:
    for line in f:
        if re.match(r"\|\s*B-\d+\s*\|", line):
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            # ID, Severity, Area, Description, Location, Found-by, Status, Resolution
            cells = (cells + [""]*8)[:8]
            # strip markdown backticks/bold for readability
            clean = [re.sub(r"[`*]", "", c) for c in cells]
            bug_rows.append([clean[0], clean[1], clean[2], clean[3], clean[6], clean[7]])
# sort by numeric id
bug_rows.sort(key=lambda r: int(re.sub(r"\D","", r[0]) or 0))
write_sheet(ws2, ["ID","Severity","Area","Description","Status","Resolution / notes"], bug_rows,
            [7,12,22,60,16,60], status_col=4)

# ============ SHEET 3: SHIPPED THIS PASS ============
ws3 = wb.create_sheet("Shipped 2026-06")
S = [
 ["Zero-knowledge cloud encryption","Security","Shipped","Financial data AES-256 encrypted on-device (password-derived key, never sent) before sync; Firestore stores only ciphertext."],
 ["Recovery code","Security","Shipped","Wrapped-key model; unlocks data after a forgotten-password reset, re-locks under new password. Shown at signup; regenerate in Settings."],
 ["In-app account deletion","Compliance","Shipped","Settings → Delete account (re-auth → wipes Firestore + Auth). App Store 5.1.1(v)."],
 ["Biometric app lock","Security","Shipped","Face ID / Touch ID / passcode + 2-min background auto-relock; Settings toggle."],
 ["Apple Privacy Manifest","Compliance","Shipped","ios.privacyManifests — no tracking; data types declared."],
 ["Hardened auth","Security","Shipped","Removed plaintext-password store; anti-enumeration messaging."],
 ["AI provider key moved server-side","Security","Shipped","Anthropic/Vision keys no longer bundled; configurable proxy (functions/aiTips)."],
 ["Accessibility labels (start)","Accessibility","Shipped","Screen-reader labels on shared controls, tab bar, TopBar, Home, back button; 44pt targets."],
 ["Formal QA suite + CI fix","Quality","Shipped","587 tests incl. golden scenarios + precision guards; CI green (Java 21 fix for the rules job)."],
 ["B-49 assets-basis divergence","Bug fix","Fixed","All surfaces use resolveNetWorthRows → one net-worth/nest-egg basis; locked by tests."],
 ["B-52 'savings rate' ambiguity","Bug fix","Fixed","Insight = 'investing % of gross'; Analytics = 'Savings rate (of take-home)' + caption."],
 ["TestFlight build v1.0.1 (#21)","Launch","Done","Built on EAS + submitted to TestFlight 2026-06-20."],
]
write_sheet(ws3, ["Item","Area","Status","Detail"], S, [40,16,12,70], status_col=2)

# Title note on sheet 1 — add a frozen banner row? keep simple: set tab colors
ws1.sheet_properties.tabColor = "C00000"
ws2.sheet_properties.tabColor = "1F4E5F"
ws3.sheet_properties.tabColor = "38761D"

wb.save(OUT)
print("wrote", OUT)
print("backlog rows:", len(B), "| bug rows:", len(bug_rows), "| shipped rows:", len(S))
