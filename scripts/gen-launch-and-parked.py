# Generates docs/finwise-launch-and-parked.xlsx — ONE workbook for prioritization (plain English).
#   Tab 1 "Launch (v1)"      : everything targeted for the NEXT launch — process blockers + the v1 work
#                              lifted up from the backlog. Plain-English items + a lifecycle Stage.
#   Tab 2 "Future backlog"   : everything NOT in the next launch (v1.x soon / v2 future).
# Your hand-entered columns (J Rank, J Notes on Tab 1; Version on Tab 2) are READ LIVE from the existing
# file and preserved on every rebuild — this script never overwrites them.
# Usage:  .venv/bin/python scripts/gen-launch-and-parked.py
import shutil
import openpyxl
import xlsxwriter

OUT  = "/Users/palakjain/Vibe_Coding/finwise/docs/finwise-launch-and-parked.xlsx"
SNAP = "/Users/palakjain/Vibe_Coding/finwise/docs/snapshots/finwise-launch-and-parked-2026-06-26.xlsx"

# ---------- 1. READ + PRESERVE the user's hand-entered columns from the existing file ----------
# Sheet-name-agnostic: scan every sheet and capture your_rank / j_rank / j_notes by ID, wherever they live.
def read_user_cols():
    pres = {}
    try:
        wb = openpyxl.load_workbook(OUT, data_only=True)
    except FileNotFoundError:
        return pres
    for ws in wb.worksheets:
        hrow, hm = None, {}
        for r in ws.iter_rows(min_row=1, max_row=6):
            if r and r[0].value == "ID":
                hrow = r[0].row
                hm = {str(c.value).strip().lower(): c.column for c in r if c.value}
                break
        if not hrow:
            continue
        def col(sub):
            for k, v in hm.items():
                if sub in k: return v
            return None
        c_yr, c_jr, c_jn = col("your rank"), col("j rank"), col("j notes")
        for row in ws.iter_rows(min_row=hrow + 1):
            idv = row[0].value
            if not idv or not str(idv)[:2] in ("L-", "P-"): continue
            g = lambda c: (ws.cell(row=row[0].row, column=c).value if c else None)
            cur = pres.setdefault(str(idv), {})
            for key, c in (("your_rank", c_yr), ("j_rank", c_jr), ("j_notes", c_jn)):
                if c and g(c) not in (None, ""):
                    cur[key] = g(c)
    return pres

USER = read_user_cols()
def uL(pid, k): return (USER.get(pid) or {}).get(k)

# ---------- 2. WORKBOOK STYLES ----------
wb = xlsxwriter.Workbook(OUT, {"in_memory": True})
HDR = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#1F4E5F", "border": 1,
                     "border_color": "#CCCCCC", "valign": "vcenter", "text_wrap": True, "font_size": 11})
TITLE = wb.add_format({"bold": True, "font_size": 13, "font_color": "#1F4E5F"})
NOTE  = wb.add_format({"italic": True, "font_size": 10, "font_color": "#555555", "text_wrap": True, "valign": "top"})
def fmt(bg=None, bold=False, italic=False):
    d = {"border": 1, "border_color": "#DDDDDD", "valign": "top", "text_wrap": True, "font_size": 10}
    if bg: d["bg_color"] = bg
    if bold: d["bold"] = True
    if italic: d["italic"] = True
    return wb.add_format(d)
BASE = fmt()
USERCOL = fmt("#FFF7DE")          # columns you fill in (pale yellow)
STAGE_BG = {"Plan": "#EAEAEA", "Design": "#E6D7F2", "Build": "#FCE5CD", "Test": "#D0E0E3",
            "Refine": "#FFF2CC", "Ready": "#D9EAD3", "Done": "#B6D7A8"}
STAGE_FMT = {k: fmt(v, bold=True) for k, v in STAGE_BG.items()}

def write_sheet(name, intro, headers, rows, widths, stage_col, user_cols, tab_color):
    ws = wb.add_worksheet(name)
    ws.set_tab_color(tab_color)
    ws.merge_range(0, 0, 0, len(headers) - 1, name, TITLE)
    ws.merge_range(1, 0, 2, len(headers) - 1, intro, NOTE)
    ws.set_row(1, 28)
    HROW = 3
    ws.freeze_panes(HROW + 1, 0)
    for c, (h, w) in enumerate(zip(headers, widths)):
        ws.set_column(c, c, w)
        ws.write(HROW, c, h, HDR)
    ws.set_row(HROW, 30)
    for r, row in enumerate(rows, HROW + 1):
        for c, val in enumerate(row):
            f = BASE
            if c == stage_col and val in STAGE_FMT:
                f = STAGE_FMT[val]
            elif c in user_cols:
                f = USERCOL
            ws.write(r, c, "" if val is None else val, f)
    ws.autofilter(HROW, 0, HROW + len(rows), len(headers) - 1)

# ---------- 3. TAB 1 DATA — Launch (v1) ----------
# Process / launch items, in critical-path order. (id, phase, item, owner, type, stage, status, note)
LAUNCH = [
 ("L-1","Data","Switch to a paid, properly-licensed source for stock & fund prices (the free Yahoo feed isn't legal for a paid app).","You→Claude","Blocker","Plan","Open — the long pole","You pick a provider (Tiingo is cheapest, ~$10–50/mo) and get a key; then I do a small swap. This is a legal must-fix."),
 ("L-15","Accuracy","Make a spending category entered as a percentage turn into the SAME dollar amount everywhere it's used.","Claude","Accuracy","Done","Done — verified in code 2026-06-27","Confirmed: % spending resolves against the take-home base everywhere (pinned by budget tests)."),
 ("L-16","Accuracy","Make every screen show the SAME number for the same thing (take-home pay can't differ screen to screen).","Claude","Accuracy","Done","Done 2026-06-25","Locked in by an automatic check. A wrong number = lost trust."),
 ("L-17","Accuracy","Confirm no money figure shown anywhere is wrong.","Claude","Important","Test","Ongoing — none currently open","Re-check before each submission."),
 ("L-2","Security","Lock down the cloud database so each person can only read their own data.","You","Blocker","Done","Done 2026-06-17 (the old checklist was stale)","Already deployed; just confirm it's active in the console."),
 ("L-19","Compliance","Write the note to Apple's reviewer explaining the app is an educational planner (no money movement).","You","Important","Refine","Draft exists","Reduces the risk of rejection under finance-app rules."),
 ("L-5","Listing","Enter the App Store description, keywords, subtitle, and promo text.","You","Blocker","Refine","Drafted — needs entering","The wording is written; paste it into App Store Connect."),
 ("L-8","Compliance","Fill in Apple's privacy questionnaire so it matches what the app actually collects.","You","Blocker","Plan","Open","Declare email, name, financial info, crash data — all to run the app, none for tracking. Apple cross-checks."),
 ("L-6","Listing","Set the app's category (Finance) and age rating (4+).","You","Blocker","Plan","Open (quick)","Short questionnaires in App Store Connect."),
 ("L-9","Compliance","Create a test login with sample data for Apple's reviewer and include it in the submission.","You","Blocker","Plan","Open","Without it the reviewer can't get past sign-in and will reject."),
 ("L-7","Listing","Publish a live Privacy Policy page and a Support page, and link them.","You","Blocker","Build","Part done — privacy page needs the new wording re-published; support page not made yet","The privacy page is live but its text changed (see L-11)."),
 ("L-11","Compliance","Re-publish the privacy page with the new wording: financial data and receipts never leave your device.","You","Blocker","Refine","Wording done in the app; the public page isn't re-published","The promise is already true in the app; the public page must say the same."),
 ("L-20","Positioning","Decide to launch in the US only for now.","You","Decision","Plan","Likely already (US-dollars only)","Confirm; don't market it globally yet."),
 ("L-21","Positioning","Decide to track 'how many people finish setup' as the #1 success measure.","You","Decision","Plan","Decision needed","Needs basic usage tracking (see the analytics item below)."),
 ("L-13","Compliance","Show a 'this is an estimate, not advice' note on every forecast or 'on-track' verdict.","Claude","Important","Done","Done","Added everywhere; a guard test stops it being removed."),
 ("L-14","Reliability","Make sure the app handles no-internet gracefully instead of crashing.","Claude","Important","Done","Done (verified)","Already safe; proven by offline tests."),
 ("L-3","Build","Build the real app installer (build #35) and check receipt-scanning + secure storage on a real phone.","You (EAS)","Blocker","Build","Blocked on the free build allowance (~Jul 1); the code is ready","Everything's in and tested; push-button once the allowance resets."),
 ("L-23","Data","Confirm prices load from the new paid source (not Yahoo).","You","Blocker","Plan","Waiting on the vendor swap (L-1)",""),
 ("L-12","Reliability","Confirm a crash report actually reaches the dashboard on the new build.","You","Blocker","Test","Wired; verify on the build","Press the test button in Settings and check it lands."),
 ("L-10","Compliance","Test that 'delete my account' fully wipes your data on a real build.","You","Blocker","Test","Open (needs build #35)","Apple requires this; confirm it removes cloud data and login."),
 ("L-22","Testing","Walk through setup as each user type (student, variable income, professional, retiree) and check the numbers add up.","You","Blocker","Test","Open (needs build #35)","Use the user guide as the script."),
 ("L-24","Testing","Test sign up → log out → recover account → log back in.","You","Blocker","Test","Open",""),
 ("L-18","Testing","Run the automated tap-through tests on a real phone (sign-up, basic screens, add-sheet, net-worth, cash-flow).","You","Important","Test","Open (needs build)","These scripts have never run on a real device, so some may need small fixes the first time."),
 ("L-25","Testing","Hands-on test on a real phone: gestures, keyboard, screen edges, large text.","You","Recommended","Test","Open (needs build)",""),
 ("L-4","Listing","Take the App Store screenshots on a real iPhone.","You","Blocker","Build","Open (needs a device build)","8 framed shots with captions are already planned out."),
 ("L-26","Submit","Submit the build to TestFlight and then to App Review.","You","Blocker","Plan","Final step","Submitting doesn't use the build allowance."),
]

# ---------- 4. BACKLOG DATA (plain English) — from the 74-item rewrite ----------
# id -> (stage, plain_item, whats_left, plain_note)
P_TXT = {
 "P-1": ("Plan","Connect the app to your bank and brokerage so balances and transactions pull in automatically.","Not started","Biggest convenience and stickiness win, but a large build needing a secure server."),
 "P-2": ("Plan","Automatically figure out dividends from a stock ticker instead of you typing them in.","Not started","Saves manual entry for people who hold dividend-paying stocks."),
 "P-3": ("Plan","Suggest stock tickers as you type and cache prices so they load faster.","Not started","Nice polish; the core performance views already work."),
 "P-4": ("Plan","Turn on snap-a-photo receipt scanning and improve how it reads the totals.","Not started","Nearly ready; needs a real device build to switch on and test."),
 "P-5": ("Plan","Catch bad rows when you import a spreadsheet so imports don't break things.","Not started","Protects against messy files corrupting your data."),
 "P-6": ("Build","Turn your saved history into smart nudges like 'gas spending up 15% this month'.","A basic 9-rule version works; the richer history-driven version isn't built.","Makes the app feel personal and proactive."),
 "P-7": ("Plan","Put a consistent tap-to-explain button on every AI suggestion so you see the reasoning.","Not started","Builds trust; explanations exist but aren't uniform yet."),
 "P-8": ("Plan","Ask the app plain questions like 'can I afford a $600k house?' and get an answer.","Not started","A standout feature that sets the app apart."),
 "P-9": ("Plan","Compare life choices like buying a home, having a child, or a job change side by side.","Not started","Helps users weigh big decisions."),
 "P-10": ("Build","Show retirees whether their money will last as they spend it down.","A basic spend-down view exists; deeper retiree-focused views aren't built.","Answers the top question retirees have."),
 "P-11": ("Plan","Plan smart, tax-savvy withdrawals in retirement to keep more of your money.","Not started","Valuable but sensitive; needs careful, accurate tax handling."),
 "P-12": ("Plan","Assume your savings grow as your pay rises over the years.","Not started","Makes retirement projections more realistic."),
 "P-13": ("Plan","Support married filers and other countries' tax and account rules.","Not started","Today it only handles single US filers; needed to expand abroad."),
 "P-14": ("Plan","Let couples and households share goals while keeping personal views too.","Partner invites are partly built.","Needed for shared finances."),
 "P-15": ("Plan","Add saving-for-college planning.","Not started","Fills a gap for families with kids."),
 "P-16": ("Plan","Add life and long-term-care insurance planning.","Not started","Fills a life-stage gap many users hit."),
 "P-17": ("Plan","Add estate planning like naming who inherits and tracking key documents.","Not started","Rounds out planning for later life stages."),
 "P-18": ("Plan","Add a job-loss mode showing how long your cash lasts and a survival budget.","Not started","Helps people through a stressful transition."),
 "P-19": ("Plan","Add a mode for gig and freelance income that smooths uneven pay and sets aside taxes.","Not started","Serves the growing self-employed audience."),
 "P-20": ("Build","Let users add more details later to sharpen their plan, prompted step by step.","A simple checklist exists; the guided 'add more to sharpen' flow isn't built.","Improves plan accuracy without overwhelming people upfront."),
 "P-21": ("Plan","Let partners securely share one encrypted account through an invite code.","Not started","Today two logins can't share the same encrypted data; needed for couples."),
 "P-22": ("Plan","Add two-step login for stronger account security.","Not started","App lock already exists; this is the next security layer."),
 "P-23": ("Plan","Scan the app's building blocks for known security holes and fix the risky ones.","Not started","A high-priority safety check before launch."),
 "P-24": ("Plan","Add automated code scanning that flags security flaws as code is written.","Not started","Catches problems early; lower priority."),
 "P-25": ("Plan","Scan the finished app file for insecure storage or data-transfer issues.","Not started","Extra security check; lower priority."),
 "P-26": ("Plan","Make sure all data fetches use secure, encrypted connections.","Not started","Protects data in transit; lower priority."),
 "P-27": ("Plan","Require a fingerprint or face check to confirm the most important actions.","Not started","App unlock already uses biometrics; this adds per-action confirmation."),
 "P-28": ("Plan","Make every screen work with the screen reader for blind and low-vision users.","Not started","The single biggest accessibility gap (for users with disabilities)."),
 "P-29": ("Plan","Make buttons big enough to tap reliably for everyone.","Not started","An accessibility basic (for users with disabilities); high priority."),
 "P-30": ("Plan","Add dark mode and a high-contrast theme.","Not started","A comfort and readability option; lower priority."),
 "P-31": ("Plan","Check that all text and charts have enough color contrast to be readable.","Not started","An accessibility review (for users with disabilities); medium priority."),
 "P-32": ("Plan","Tone down animations for users who prefer less motion.","Not started","A small comfort improvement; low priority."),
 "P-33": ("Done","Add a tap to hide or blur your balances for privacy.","Fixed 2026-06-27 (B-90): all 36 leaking displays now masked, locked by a guard test. Re-verify on build #36.","Lets you check the app in public without showing your money."),
 "P-34": ("Plan","Show how many months your cash savings could cover if income stopped.","Not started","A reassuring 'safety cushion' number; overlaps the job-loss mode."),
 "P-35": ("Plan","Add a Home card showing how your net worth has grown over time.","Not started","A motivating progress view like 'net worth up $40k in 6 months'."),
 "P-36": ("Build","Redesign the top bar and bottom menu to match the newer look.","The newer style exists elsewhere; the header and menu still use older styling.","Visual consistency across the app."),
 "P-37": ("Plan","Refresh the older detail screens like budget, transactions, settings, and rewards.","Not started","The new home tiles still open older-looking screens."),
 "P-38": ("Plan","Make total spending and the category breakdown always add up to the same number.","Not started","Prevents confusing double-counting between summaries and budget."),
 "P-39": ("Plan","Add loading placeholders and clear offline / syncing states.","Not started","Makes the app feel smoother while data loads; medium priority."),
 "P-40": ("Plan","Add an app-wide banner showing when you're offline or syncing.","Not started","Keeps users informed about connection status; lower priority."),
 "P-41": ("Plan","Add an 'undo' option after deleting something by mistake.","Not started","Friendlier than a blocking confirmation pop-up; medium priority."),
 "P-42": ("Build","Add a reusable 'what's this?' tooltip and a searchable plain-language glossary.","Simple explainer dots were added; the full tooltip and glossary module may be unfinished.","Helps users understand finance terms; lower priority."),
 "P-43": ("Plan","Set one consistent rule for showing money and dates so cents don't drift.","Not started","Keeps numbers looking correct and uniform; lower priority."),
 "P-44": ("Plan","Let users pin and rearrange the cards on their Home screen.","Not started","Personalizes the home view; lower priority."),
 "P-45": ("Plan","Let users switch any chart to a plain data table.","Not started","Easier to read for some users; lower priority."),
 "P-46": ("Plan","Build a notification system with frequency limits and per-type controls.","Not started","Prevents notification overload; lower priority."),
 "P-47": ("Plan","Explain why the app needs camera or notification access before the system asks.","Not started","Increases the chance users say yes; lower priority."),
 "P-48": ("Plan","Show a visible 'your data is encrypted' cue in the app.","Not started","Reassures users their data is protected; low priority."),
 "P-49": ("Plan","Make form error messages appear at consistent moments across the app.","Not started","Small polish for a smoother experience; low priority."),
 "P-50": ("Plan","Define standard text styles for titles, body, and captions.","Not started","Keeps typography consistent; low priority."),
 "P-51": ("Plan","Check that layouts work for one-handed use and large text sizes.","Not started","A comfort and accessibility review (for users with disabilities); low priority."),
 "P-52": ("Plan","Send reminder nudges like 'you're near your dining budget'.","The plumbing exists; the reminder logic isn't built.","Drives people back into the app."),
 "P-53": ("Build","Finish the full savings-goals screen with progress, priority, and projections.","The goals tab and funding logic exist; the full visual screen isn't finished.","Helps people set and track goals; a stickiness feature."),
 "P-54": ("Plan","Add younger, motivational framing and a simpler app-wide mode.","Not started","Makes the app approachable; only one screen has a plain toggle today."),
 "P-55": ("Plan","Add more game-like rewards, mini lessons, and shareable wins.","Not started","Boosts engagement through milestones and streaks."),
 "P-56": ("Plan","Add trends, a year-in-review, shareable reports, widgets, and a weekly summary.","Not started","Turns saved history into ongoing value for users."),
 "P-57": ("Plan","Translate the app's wording into other languages, not just the numbers.","Not started","Needed to reach non-English users; currency formatting already done."),
 "P-58": ("Build","Confirm money and date formats display correctly on a real device.","The formatting logic is wired; it just needs checking on an actual phone.","Ensures numbers look right everywhere."),
 "P-60": ("Plan","Let users export their data to a file and import it back.","Not started","Gives users control of their data; medium priority."),
 "P-61": ("Plan","Safely upgrade saved data as new fields are added over time.","Not started","Prevents data problems when the app updates."),
 "P-62": ("Plan","Add tests for the core retirement and investment math.","Not started","Protects accuracy of the most important calculations; high priority."),
 "P-63": ("Plan","Review that every data fetch handles errors gracefully, not just the happy path.","Not started","Prevents crashes and blank screens; high priority."),
 "P-64": ("Plan","Add timeouts and retries so the app never hangs waiting on data.","Not started","Stops endless loading spinners; medium priority."),
 "P-65": ("Plan","Replace outdated app components with their current versions.","Not started","Keeps the app maintainable and future-proof; medium priority."),
 "P-66": ("Plan","Simplify the most tangled parts of the code, like the retirement and income screens.","Not started","Makes future changes safer and easier; lower priority."),
 "P-67": ("Plan","Remove unused and leftover old code.","Not started","Tidies up the app after launch; lower priority."),
 "P-68": ("Plan","Tighten up loose data typing to catch more bugs automatically.","Not started","Reduces hidden errors; lower priority."),
 "P-69": ("Plan","Require core calculations to stay well-tested before any code change ships.","Not started","Guards the most important math automatically."),
 "P-70": ("Plan","Make design consistency a required check before new code is merged.","Not started","The rules are written but not yet enforced."),
 "P-71": ("Plan","Pick and trademark a final app name, since 'FinWise' may already be taken.","Not started","A real legal blocker before public launch; 'Hatcho' is the front-runner."),
 "P-72": ("Plan","Improve overall speed and make the app work well offline.","Not started","A broad, behind-the-scenes quality improvement."),
 "P-73": ("Plan","Add usage tracking to see how many people finish setup and keep coming back.","Not started","Needed to measure success and the top launch metric."),
 "P-74": ("Plan","Test the app across many real phone models in the cloud to catch device-specific bugs.","Not started","Ensures it works on the wide range of phones people own."),
 # P-59 (crash reporting wiring) intentionally DROPPED here — it's the same work as launch item L-12.
}
P_CAT = {  # short, plain category per id
 **{f"P-{i}": "Data" for i in [1,2,3,4,5]},
 **{f"P-{i}": "Insights" for i in [6,7,8,9]},
 **{f"P-{i}": "Retire" for i in range(10,20)},
 "P-20": "Onboarding",
 **{f"P-{i}": "Security" for i in range(21,28)},
 **{f"P-{i}": "Accessibility" for i in range(28,33)},
 **{f"P-{i}": "UI/UX" for i in range(33,52)},
 **{f"P-{i}": "Growth" for i in range(52,57)},
 **{f"P-{i}": "Quality" for i in range(57,71)},
 "P-71": "Branding",
 **{f"P-{i}": "Other" for i in [72,73,74]},
}
# --- Three buckets (re-triaged from your v1 tags into must-ship / fast-follow / future) ---
# V1 = genuinely must-ship to launch. Only the legal blocker + two safety items join the L-# process list.
V1_BACKLOG = {  # id -> (Type, Owner, prefilled Decision)
 "P-71": ("Blocker","You→Claude","Before submit — legal"),
 "P-23": ("Important","Claude","Before submit — safety"),
 "P-63": ("Important","Claude","Before submit — safety"),
 "P-33": ("v1 feature","Claude","DONE"),   # already built; ships in the launch build
}
# FAST FOLLOW-UP = my recommendation for the first quick update right after launch (ordered by priority).
FAST = ["P-62","P-38","P-64","P-73","P-4","P-5","P-53","P-28","P-29","P-41","P-39","P-34","P-60"]
# Everything else in P_TXT (not V1_BACKLOG, not FAST, not the dropped P-59) → V2.

# ---------- 5. BUILD TAB 1 ROWS ----------
# When each launch item should happen (my recommendation; you can override).
LAUNCH_DECISION = {
 "L-1":"NOW","L-15":"DONE","L-16":"DONE","L-17":"NOW — verify","L-2":"DONE — confirm console",
 "L-19":"NOW","L-5":"NOW","L-8":"NOW","L-6":"NOW","L-9":"NOW","L-7":"NOW","L-11":"NOW",
 "L-20":"NOW — decide","L-21":"NOW — decide","L-13":"DONE","L-14":"DONE",
 "L-3":"AT BUILD #35","L-23":"AT BUILD #35","L-12":"AT BUILD #35","L-10":"AT BUILD #35",
 "L-22":"AT BUILD #35","L-24":"AT BUILD #35","L-18":"AT BUILD #35","L-25":"AT BUILD #35",
 "L-4":"AT BUILD #35","L-26":"SUBMIT (last)",
}
T1_HEAD = ["ID","Area","Item","Owner","Type","Stage","Status / what's left","Your rank","J Rank","J Notes","Decision (when)","Notes / next step"]
t1 = []
for (idv, phase, item, owner, typ, stage, status, note) in LAUNCH:
    t1.append([idv, phase, item, owner, typ, stage, status,
               uL(idv,"your_rank"), uL(idv,"j_rank"), uL(idv,"j_notes"),
               LAUNCH_DECISION.get(idv,""), note])
# append the few V1 backlog items (blocker first, then important, then the done feature)
def pnum(pid): return int(pid.split("-")[1])
v1movers = sorted(V1_BACKLOG, key=lambda p: ({"Blocker":0,"Important":1,"v1 feature":2}[V1_BACKLOG[p][0]], pnum(p)))
for pid in v1movers:
    stage, item, wl, note = P_TXT[pid]
    typ, owner, decision = V1_BACKLOG[pid]
    status = "Done — verify" if stage == "Done" else (wl if wl != "Not started" else "Not started")
    t1.append([pid, P_CAT.get(pid,"Backlog"), item, owner, typ, stage, status,
               "", uL(pid,"j_rank"), uL(pid,"j_notes"), decision, note])

write_sheet("V1 — Next launch",
  "EVERYTHING ON THIS TAB IS V1 — the genuine must-ship list for the next launch, nothing optional. It's the ship "
  "process (L-#, in critical-path order) plus the only backlog items that truly can't wait: the final app name "
  "(legal), two safety items, and one feature that's already built. Type: 'Blocker' = can't ship without it · "
  "'Important' = strongly recommended before submit · 'Decision'/'Recommended' as labelled. Stage = where it is "
  "(Plan→Design→Build→Test→Refine→Ready, or Done). Yellow columns are yours (J Rank / J Notes / Decision); 'Your "
  "rank' is my suggested order. Things I'd do right after launch are on the 'Fast follow-up' tab; everything else "
  "is on 'V2 — Future'.",
  T1_HEAD, t1, [6, 12, 50, 11, 11, 8, 26, 8, 7, 14, 16, 46],
  stage_col=5, user_cols={7,8,9,10}, tab_color="#C00000")

# ---------- 6. BACKLOG TABS — Fast follow-up + V2 (everything NOT in the next launch) ----------
BL_HEAD = ["ID","Category","Item","Stage","What's left","Your notes","Why it matters"]
def backlog_row(pid):
    stage, item, wl, note = P_TXT[pid]
    return [pid, P_CAT.get(pid,""), item, stage, ("" if wl == "Not started" else wl), "", note]

# Fast follow-up — my recommended first update after launch, in priority order.
fast_rows = [backlog_row(pid) for pid in FAST if pid in P_TXT]
write_sheet("Fast follow-up",
  "MY RECOMMENDATION for the first quick update right after launch — high-value, lower-risk improvements that don't "
  "block the launch but I'd ship soon after. Listed in priority order (top = first). These were pulled out of the "
  "backlog so they don't get lost behind the bigger V2 work. Want any of these IN the launch instead? Tell me and "
  "I'll move it to V1. Want one pushed further out? It goes to V2.",
  BL_HEAD, fast_rows, [6, 13, 52, 8, 34, 18, 52],
  stage_col=3, user_cols={5}, tab_color="#E69138")

# V2 — everything else, in id order.
v2_rows = [backlog_row(f"P-{i}") for i in range(1, 75)
           if f"P-{i}" in P_TXT and f"P-{i}" not in V1_BACKLOG and f"P-{i}" not in FAST]
write_sheet("V2 — Future",
  "EVERYTHING ON THIS TAB IS V2 — a later, bigger release down the road (bank linking, the planning copilot, "
  "couples, more retirement modelling, translations, design polish, etc.). If it's not on the V1 or Fast follow-up "
  "tabs, it's here. Stage shows how far along each is; 'What's left' spells out what's unfinished for the part-done "
  "ones. The yellow 'Your notes' column is yours. To pull anything forward, tell me which tab to move it to. "
  "(The crash-reporting row was removed — it's the same work as launch item L-12.)",
  BL_HEAD, v2_rows, [6, 13, 52, 8, 34, 18, 52],
  stage_col=3, user_cols={5}, tab_color="#1F4E5F")

wb.close()
shutil.copyfile(OUT, SNAP)
print("wrote", OUT)
print("snapshot", SNAP)
print("V1 rows:", len(t1), "| Fast follow-up rows:", len(fast_rows), "| V2 rows:", len(v2_rows), "| dropped: P-59")
print("preserved user values for", len(USER), "rows")
