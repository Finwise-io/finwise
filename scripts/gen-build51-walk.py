#!/usr/bin/env python3
"""The Build-51 founder walk sheet — what changed since Build 50, in the order you'd tap it.

One row per thing to look at, phrased as what YOU do and what you should see. Yellow columns are
yours: Pass/Fail and a comment. Nothing in here is a code instruction — it is a walk.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUT = 'docs/FCC-core-55-70/build51-founder-walk-2026-08-11.xlsx'
HEAD = PatternFill('solid', fgColor='085041')
YOURS = PatternFill('solid', fgColor='FFF7D6')
BAND = PatternFill('solid', fgColor='DFF2E9')
HF = Font(color='FFFFFF', bold=True, size=11)

WALK = [
    # (area, what you do, what you should see)
    ('Net worth — the top', 'Open the Net worth tab',
     'The title bar is a GREEN BANNER running the full width of the screen, reading "YOUR NET WORTH · AUG 11, 2026" with the info dot. No white card around it.'),
    ('Net worth — the top', 'Look down the screen at the five section titles',
     'Hero, Your retirement plan, What you own, What you owe and Emergency cushion all wear the SAME green banner, edge to edge. The screen reads as one instrument, not a stack of cards.'),
    ('Net worth — hero', 'Read the hero, top to bottom',
     'Own − Owe on a small line, then the one big green number, then "Change in net worth +$X · Return on cash + investments +Y%", then "since <date>". Nothing after that — no green line at the bottom.'),
    ('Net worth — hero', 'Check the two dates are different things',
     'The banner date is TODAY (when the numbers are true as of). The "since" date is when the change is measured FROM. Neither repeats the other.'),
    ('Net worth — hero', 'Tap the change line',
     'The walk sheet opens: How $X became $Y, the rows summing exactly to the ending number.'),
    ('Net worth — hero', 'Tap the ⓘ on the banner',
     'Plain-English explanation of how the change and the return are worked out.'),
    ('Net worth — what you own', 'Look at the bar under WHAT YOU OWN',
     'A thin stacked bar (no donut) with a legend naming every slice and its percent — Real estate 55%, Stocks / ETFs 43%, Cash 1%, Bonds & CDs 1%. The percentages total 100.'),
    ('Net worth — what you own', 'Check the ORDER of the categories',
     'Biggest first, in the bar AND in the list beneath it — the same story in the same order.'),
    ('Net worth — alignment', 'Run your eye down the right-hand edge of every number',
     'THE ONE THING TO CHECK HARDEST: every number — section totals, group totals, category totals, account amounts — sits on ONE right edge. Nothing sits slightly left of anything else.'),
    ('Net worth — what you own', 'Tap a group bar (Cash / Investments / Personal property)',
     'It collapses and expands. Closed, it says how many are behind it ("· 2 accounts", "· 2 categories", "· 1 item").'),
    ('Net worth — grouping', 'Tap "By institution"',
     'Your banks and brokerages, biggest first, each one collapsible. An institution shows only what it HOLDS — a card you owe at the same bank is NOT netted off it.'),
    ('Net worth — grouping', 'Check the Chase Visa is not inside Chase',
     'Chase shows its checking balance only. The Visa lives under What you owe, where it belongs.'),
    ('Net worth — what you owe', 'Look under WHAT YOU OWE',
     'Its own bar in the warning colours with a named legend, then the debt rows. The costliest debt carries the "pay first" tag.'),
    ('Net worth — cushion', 'Look at Emergency cushion',
     'Months, the Tight/Comfortable word, a progress bar filling against the 3–6 month guide, and the maths line ("$X cash ÷ $Y/mo essentials").'),
    ('Net worth — cushion', 'Check the cushion divides by the same cash the CASH group shows',
     'The two figures must be identical. They were different before this build ($8,838 in the group vs $8,000 in the cushion).'),
    ('Net worth — accounts', 'Tap any account row',
     'It opens that account\'s own page. Every row that shows an arrow goes somewhere.'),
    ('Net worth — what is gone', 'Look below the cushion',
     'The "This month\'s cash flow" row is GONE (your call). The "＋ Add or connect an account" button is still there.'),
    ('Net worth — first day', 'Sign in as a brand-new user (or reinstall)',
     'The whole layout stays at zero: Own $0 − Owe $0, a green $0, "Add your first account and this becomes your one live number", every group at $0, the three ways in, debts and cushion in place.'),
    ('Net worth — first day', 'Check the grouping buttons are there on day one',
     'By category / By institution both show (your call). Tapping By institution says "Your banks and brokerages will be listed here, each with what it holds" — no invented $0 banks.'),
    ('Net worth — first day', 'Tap "Connect a brokerage" on day one',
     'It opens the connect flow. It used to lead nowhere and bounce you back Home — fixed in this build.'),
    ('Net worth — banner', 'With a stale connection or an unpriced holding, look under the hero',
     'A tinted card naming each gap. Tap it: one row per gap with a button that lands on the exact cure. Fix them and the card removes itself.'),
    ('Net worth — banner', 'In that sheet, tap "Update the value" on an old property value',
     'It opens that account\'s page. It used to lead nowhere — fixed in this build.'),
    ('Data honesty', 'Look at any connected account row',
     'It reads "Connected · updated today" — NOT "updated updated today", which is what every connected row said before this build.'),
    ('Data honesty', 'Check where a CD or Treasuries account sits',
     'Under Bonds & CDs, not Cash and not Unclassified — including accounts you connected before this build.'),
    ('Data honesty', 'Check where a money-market fund sits',
     'Under Stocks / ETFs (it pays dividends), not Cash — including ones stored before this build.'),
    ('Performance tab', 'Open Performance',
     'Summary trio, the value walk in the approved order, per-account walks that expand, period chips carrying their own returns.'),
    ('Across the app', 'Turn on hide-balances (the eye in the top bar)',
     'Every money figure on Net worth masks — hero, totals, group bars, rows, the cushion maths.'),
    ('Device-only — yours', 'Turn text size up to the largest setting',
     'The dated banner title and the bar legends stay readable and do not clip. THIS CANNOT BE CHECKED BY A TEST — it needs your eyes on the phone.'),
    ('Device-only — yours', 'Turn on reduce-motion',
     'Sheets appear without animation; nothing jumps.'),
    ('Device-only — yours', 'Take a grayscale screenshot of Net worth',
     'Both bars still readable: every slice is named with its percent, so colour is never doing the work alone.'),
    ('Device-only — yours', 'Swipe to the app switcher',
     'The Net worth screen is masked in the snapshot.'),
    ('Device-only — yours', 'On the smallest phone you have, open an editor sheet',
     'The keyboard never covers the save button.'),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Build 51 walk'
heads = ['#', 'Area', 'What you do', 'What you should see', 'Pass / Fail', 'Your comment']
ws.append(heads)
for c in range(1, len(heads) + 1):
    ws.cell(1, c).fill = HEAD
    ws.cell(1, c).font = HF
ws.row_dimensions[1].height = 24
for i, (area, do, see) in enumerate(WALK, 1):
    ws.append([i, area, do, see, '', ''])
for col, w in zip('ABCDEF', [5, 24, 46, 88, 12, 40]):
    ws.column_dimensions[col].width = w
for r in range(2, ws.max_row + 1):
    for c in range(1, 7):
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    for c in (5, 6):
        ws.cell(r, c).fill = YOURS
    if 'Device-only' in str(ws.cell(r, 2).value):
        ws.cell(r, 2).fill = BAND
ws.freeze_panes = 'A2'

# ── the context tab ──
ws2 = wb.create_sheet('What is in this build')
rows = [
    ['Build', 'v1.1.0 BUILD 51, cut 2026-08-11 from commit 8e09374 on branch taxonomy-v1.0.7. NOTE: not 50 — build 50 is the Aug-10 build (commit 58bd1d1) and has none of this work in it. If the Net worth title still sits on a white card, you are on 50.'],
    ['Since', 'Build 50 (2026-08-10, commit 58bd1d1) — 22 commits'],
    ['The headline', 'Net worth rebuilt to the Claude Design "Quiet Instrument" handoff: one flat ledger where every number shares one right edge, the composition bar in place of the donut, collapsible institutions, and one green banner titling all five sections.'],
    ['Your decisions built', 'Bar replaces donut · debt bar · cushion progress bar · categories biggest first · cash-flow row deleted · grouping buttons on day one · real account names kept · white hero card kept · today\'s date on the hero title.'],
    ['Defects fixed on the way', 'Two buttons that led to screens that do not exist · connected accounts landing in Unclassified · two different cash numbers on one screen · "updated updated today" on every connected row · a figure that could not be computed printing as a real-looking $0 · an audit fixture that had been hiding all debts from the audit.'],
    ['Gates at the cut', '1,492 tests green · types clean · UI-test gate passing · 4 journey suites (64 tests) green · open-gaps ledger 9 of 9 closed with code evidence.'],
    ['Known and open', 'Live price provider still unselected (your call). Debt bar shows whole percentages (99% / 1%) not the handoff\'s 98.6% / 1.4% — my call, reversible. Mocks for the other 19 flow screens drawn but not reviewed by you.'],
    ['Not verified by anything but your eyes', 'How the dated banner title and the bar legends wrap at the largest text size on the smallest phone.'],
]
ws2.append(['Item', 'Detail'])
for c in (1, 2):
    ws2.cell(1, c).fill = HEAD
    ws2.cell(1, c).font = HF
for r in rows:
    ws2.append(r)
ws2.column_dimensions['A'].width = 26
ws2.column_dimensions['B'].width = 120
for r in range(2, ws2.max_row + 1):
    for c in (1, 2):
        ws2.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
ws2.freeze_panes = 'A2'

wb.save(OUT)
print('wrote', OUT, '—', ws.max_row - 1, 'walk rows')
