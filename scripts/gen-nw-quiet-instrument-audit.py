#!/usr/bin/env python3
"""Builds the Net-worth Quiet-Instrument audit workbook (founder mock-match standing order).

Tabs
  1 Read me            — what this rebuild is, and how the audit was run
  2 Founder notes      — every line of the 2026-08-10 NW-screen notes, with where it landed
  3 Mock-vs-build      — the string/token diff, state by state, with the coverage line
  4 Defects found      — problems the audit turned up (beyond the redesign) and their fix
  5 Open founder calls — the decisions I did NOT make on my own
  6 Appearance audit   — the rendered-style checks (standing order step 5)
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUT = 'mockups/NW-QuietInstrument-Aug-10/NW-quiet-instrument-audit-2026-08-10.xlsx'
HEAD = PatternFill('solid', fgColor='085041')
BAND = PatternFill('solid', fgColor='DFF2E9')
WARN = PatternFill('solid', fgColor='FAEEDA')
ASK = PatternFill('solid', fgColor='E6F1FB')
HF = Font(color='FFFFFF', bold=True, size=11)


def sheet(wb, title, headers, rows, widths, fills=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).fill = HEAD
        ws.cell(1, c).font = HF
        ws.cell(1, c).alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 22
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
        if fills:
            f = fills(ws.cell(r, 1).value, [ws.cell(r, c).value for c in range(1, len(headers) + 1)])
            if f:
                for c in range(1, len(headers) + 1):
                    ws.cell(r, c).fill = f
    ws.freeze_panes = 'A2'
    return ws


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── 1 Read me ────────────────────────────────────────────────────────────────
sheet(wb, 'Read me', ['Item', 'What it says'], [
    ['What this is', 'The Net worth main screen rebuilt to the Claude Design "Quiet Instrument" handoff you sent on 2026-08-10, reconciled with your NW-screen notes of the same day and with the decisions already approved for this app.'],
    ['The mocks', 'mockups/NW-QuietInstrument-Aug-10/ — v1 is the handoff drawn exactly as received; v2 is the BUILD SPEC (what I actually built) with the differences and the reasons written on it. The handoff files themselves are archived in the same folder.'],
    ['How the audit was run', 'Your standing order, all five steps: (1) pull every visible string from the mock in order; (2) render the screen in a test with your data shape and dump every visible string in order; (3) diff the two; (4) check order, nesting, alignment and data reality; (5) pull the RENDERED styles and check colours, sizes, the shared right edge, spacing and the donut.'],
    ['Coverage — ROUND 2 (after your six decisions)', 'With data: 142 of 150 mock lines identical and in the same order. First day: 98 of 103. Combined 240 of 253, and every one of the 13 unmatched lines is accounted for: 3 are the info dot (the extractor reads the mock\'s dot as a letter and the screen\'s as decoration — both screens show it), 2 are the retirement figures the engine computes for the test person rather than the mock\'s illustration, and 2 are the tab bar, which belongs to the app shell rather than this screen. NO unexplained differences remain.'],
    ['Coverage — round 1 (before your decisions)', 'With data 124 of 153, first day 92 of 101 — five real differences and one gap, which became the six questions you answered.'],
    ['Gates', 'Full suite 1490 passing (was 1464 before this work) · types clean · UI-test gate passing. 26 new tests, each pinning one thing this work changed.'],
    ['What I did NOT do', 'I did not cut a build.'],
], [24, 132])

# ── 1b Round 2: the decisions ────────────────────────────────────────────────
sheet(wb, 'Round 2 — your decisions', ['#', 'Your decision', 'What changed in the app', 'Proof'], [
    ['Q1', 'Keep the real account name',
     'No code change — the cash row still reads "Vanguard Brokerage" with "cash in this account" beneath it, so the row you tap is the account you open. The MOCK was corrected to match your decision.',
     'Mock v3 updated; the screen dump was already this way'],
    ['Q2', 'Biggest first',
     'Categories now sort largest-to-smallest, in the bar AND in the list, so the picture and the rows tell the same story in the same order. Inside Investments: Stocks / ETFs $348,495 above Bonds & CDs $5,819. This SUPERSEDES the older "order by liquidity, cash first" rule from the pre-48 audit — noted in the code and in the changelog so nobody re-applies the old rule by accident.',
     'Test: "categories read BIGGEST FIRST inside a group, matching the bar above them"'],
    ['Q3', 'Keep add or connect. Delete cashflow.',
     'The "This month\'s cash flow" row is deleted from this screen, along with the code that computed it. The add-or-connect button stays. The emergency cushion keeps its own door into Cash flow for the case where it needs one.',
     'Test: "the cash-flow glance is gone from this screen (founder Q3)"'],
    ['Q4', 'Yes — the buttons give the user a view of what they can do',
     'By category / By institution now show on the first-day screen, rendered from the SAME control the full screen uses so the two can never drift. By institution at zero says what it will hold ("Your banks and brokerages will be listed here, each with what it holds") instead of drawing three $0 banks that do not exist.',
     'Test: "first day: the grouping buttons are there, and By institution says what it will hold"'],
    ['Q5', 'Keep the white hero card with green band',
     'No change.', 'Unchanged since round 1'],
    ['Q6', 'Add both. Bar replaces donut.',
     'The donut is REMOVED (component and its styles deleted). What you own now carries one thin stacked bar with a named legend; what you owe carries the same bar shape in the warning colours; the emergency cushion carries a progress bar filling against the 3-6 month guide its own note names. Percentages always total 100 — the rounding remainder goes to the largest slice, and a slice too small to reach 1% still gets a visible mark and reads "<1%" rather than vanishing.',
     '7 arithmetic tests + "the composition bar names every slice with its percent" + "the cushion carries its own progress bar"'],
    ['—', 'Something to look at (my call, reversible)',
     'The debt bar reads "Home mortgage 99% · Chase Visa 1%". The handoff drew 98.6% / 1.4%. I used whole percentages on BOTH bars so the two read the same way on one screen. Say the word for one decimal place on debts.',
     'Mock v3 header note'],
    ['—', 'One number, one picture',
     'While wiring the cushion bar I found it would have been drawn from 1.96 months while the label read 2.0. The bar now uses the same rounded figure the person can see — a chart must never be drawn from a number that is not on screen.',
     'Test asserts the bar width matches the shown 2.0 months'],
], [6, 44, 92, 52], fills=lambda k, row: BAND)

# ── 2 Founder notes ──────────────────────────────────────────────────────────
sheet(wb, 'Founder notes', ['#', 'Your note (2026-08-10)', 'What I did', 'Where'], [
    [1, 'Hero box: title still show only info dot', 'The band now reads "YOUR NET WORTH" with the info dot and nothing else — the date moved off it entirely.', 'NetWorthScreen hero'],
    [2, 'Box has a green line at the bottom — delete it', 'Deleted. That was the trend sparkline; the hero now ends at the since line, as you asked.', 'NetWorthScreen hero'],
    [3, 'Change text: "Change in NW <$0/+$x/-$x> · Return on cash + investments <%>", green for zero or positive, amber for negative',
     'Built, with the words spelled out in full: "Change in net worth +$2,110 · Return on cash + investments +0.5%". Zero and up are green; only a fall is amber. I wrote "net worth" rather than "NW" — your own no-initialisms rule.', 'NetWorthScreen change line'],
    [4, 'Since date with year, not in bold', 'Built — "since Aug 3, 2026", regular weight, its own line.', 'NetWorthScreen since line'],
    [5, 'Hero box should end after this', 'It does. Nothing follows the since line inside the hero.', 'NetWorthScreen hero'],
    [6, 'Day 1 hero: Own $0 − Owe $0 / $0 in green / "Add your first account and this becomes your one live number ›"',
     'Built, word for word, and the number is green.', 'NetWorthScreen first-day'],
    [7, 'Banner box background is not white; inline under the hero; tapping opens a sheet with a fix button per gap',
     'Already built that way and kept — the banner is an amber-tinted card directly under the hero, and it disappears by itself when the gap closes. Its "Update the value" button was landing on a route that does not exist; fixed.', 'DataGapsBanner · gaps engine'],
    [8, 'Retirement plan box, with the day-1 sample', 'Kept as approved, now with its info dot and arrow. Day one shows the one app-wide sample, identical to Home.', 'NetWorthScreen retirement card'],
    [9, 'Buttons left justified: By Category / By Institution', 'Built — the two buttons now sit on the ledger\'s left edge instead of centred.', 'NetWorthScreen grouping pills'],
    [10, 'WHAT YOU OWN: title on a dark green band, total right justified, table format so every $ is right justified on one line',
     'Built, and this is the biggest structural change: the own/owe lists are now one flat table. Every number on the screen — section totals, group totals, row amounts — lands on ONE right edge, because every row reserves the same arrow column. This is the "totals sit slightly left" problem fixed at the root rather than nudged.', 'NetWorthScreen ledger · SectionBand'],
    [11, 'Donut: put % and labels on the donut; colours a colourblind person can read',
     'Built — the two big slices carry their name and percentage beside the ring with a colour swatch, and slices too small to label without collision are named on one line beneath. Colours are the already-validated set, and no slice relies on colour alone.', 'NetWorthScreen ClassDonut'],
], [5, 52, 78, 30])

# ── 3 Mock vs build ──────────────────────────────────────────────────────────
sheet(wb, 'Mock-vs-build', ['State', 'Kind', 'The mock says', 'The screen shows', 'Verdict'], [
    ['A (with data)', 'SAME', 'Hero: Own $813,152 − Owe $418,000 · $395,152 · ▲ Change in net worth +$2,110 · Return on cash + investments +0.5% › · since Aug 3, 2026', 'Identical, in that order', 'Match'],
    ['A', 'SAME', 'WHAT YOU OWN $813,152 · CASH $8,838 · INVESTMENTS $354,314 · PERSONAL PROPERTY $450,000', 'Identical', 'Match — and every figure reconciles: 8,000+838+348,495+5,819+450,000'],
    ['A', 'SAME', 'WHAT YOU OWE −$418,000 · Home mortgage −$412,000 · Chase Visa pay first −$6,000', 'Identical', 'Match'],
    ['A', 'SAME', 'EMERGENCY CUSHION 2.0 months Tight ⚠ · $8,838 ÷ $4,500/mo', 'Identical', 'Match (was 1.8 months before this pass — see Defects, D3)'],
    ['A', 'DIFFERENT', 'Cash row reads "Vanguard — sweep cash"', 'Reads "Vanguard Brokerage" with "cash in this account" beneath it', 'Your call — the screen names the real account and says which part of it is cash; the mock uses a shorter made-up name. See Open calls, Q1.'],
    ['A', 'DIFFERENT', 'Inside INVESTMENTS: Stocks / ETFs first, then Bonds & CDs (biggest first)', 'Bonds & CDs first, then Stocks / ETFs', 'Your call — the build follows the approved rule "order classes by liquidity, cash first". See Open calls, Q2.'],
    ['A', 'DIFFERENT', 'Class rows drawn open, personal property drawn closed', 'Class rows start closed, the three groups start open', 'Follows the approved rules (a class row opens on tap; groups remember what you leave open). The mock drew one particular moment. No change made.'],
    ['A', 'DIFFERENT', 'Retirement sentence: ~$1,412,000 / ~$890,000', 'The engine\'s figures for the test person', 'Not a copy difference — same sentence, real numbers.'],
    ['A', 'EXTRA', '(not in the mock)', '"This month\'s cash flow" row and the "＋ Add or connect an account" button', 'Kept deliberately — both are approved features and the mock simply stops above them. Say the word and they go. See Open calls, Q3.'],
    ['C (first day)', 'SAME', 'The full layout at zero: Own $0 − Owe $0, $0 hero, the line about your first account, the one app-wide sample, every group at $0, the three ways in, debts, cushion', 'Identical', 'Match — 92 of 101 tokens'],
    ['C', 'MISSING', 'The By category / By institution buttons appear on the first-day screen', 'They are not rendered until you have an account', 'Your call — on an empty screen the second button would change nothing, which is a dead control on the first screen a person ever sees. See Open calls, Q4.'],
    ['D (banner)', 'SAME', 'Tinted card under the hero, gap count, one line per gap, opens the fix sheet', 'Identical', 'Match'],
    ['E (change walk)', 'SAME', 'HOW $X BECAME $Y, the same rows in the same order as the Performance walk, ending line, rows sum by construction', 'Identical — untouched by this rebuild', 'Match'],
    ['F (fix sheet)', 'SAME', 'One row per gap, what we do meanwhile, a button that lands on the cure', 'Identical, and one button that used to land nowhere now works', 'Match (see Defects, D1)'],
], [14, 12, 58, 52, 62],
      fills=lambda k, row: WARN if row[1] in ('DIFFERENT', 'MISSING') else (ASK if row[1] == 'EXTRA' else None))

# ── 4 Defects found ──────────────────────────────────────────────────────────
sheet(wb, 'Defects found', ['#', 'What was wrong', 'Who would have hit it', 'Fixed', 'Pinned by'], [
    ['D1', 'Two buttons pointed at screens that do not exist. The first-day "Connect a brokerage ›" pushed /connect-account (the screen is /connect), and the missing-data sheet\'s "Update the value ›" pushed /(tabs)/networth (this tab\'s route is /(tabs)/analytics). Both bounced the person back to Home.',
     'Anyone starting on day one, and anyone with a property value older than a year', 'Yes — both now open the right screen',
     'A new test walks every route the app pushes and fails if any names a screen that does not exist. I put both bugs back to prove it catches them.'],
    ['D2', 'A connected account that sends a balance but no holdings was dumped whole into "Unclassified" — even a CD/Treasuries account whose type we know exactly. Same class of problem as the CD-under-Cash one you found.',
     'Anyone with a connected fixed-income or savings account — it showed as Unclassified in the donut and the list', 'Yes — an account keeps its own class; only a wrapper whose contents are genuinely unknown stays Unclassified',
     '3 engine tests (bonds account, savings account, and a bare brokerage that must STAY unclassified)'],
    ['D3', 'The screen printed two different cash numbers an inch apart: the CASH group counted a connected brokerage\'s sweep balance, the emergency cushion did not. It read "CASH $8,838" above "$8,000 cash ÷ $4,500/mo" — 1.8 months instead of 2.0.',
     'Anyone with sweep cash inside a connected brokerage', 'Yes — both read the same breakdown now',
     'An agreement test asserting the group total and the cushion numerator are the same number'],
    ['D4', 'Every connected account row read "Connected · updated updated today". The freshness label already contained the word "updated" and each caller added it again — four places across three screens, including the spoken labels for a screen reader.',
     'Everyone with a connected account', 'Yes — fixed at the source, so all four read correctly',
     'Existing freshness tests still pass; the wording now appears correctly in the screen dump'],
    ['D5', 'A figure the app could not compute printed as "$0" — a real-looking balance. That is how a Net worth screen whose debts failed to load showed a confident "$0" instead of admitting it had nothing.',
     'Anyone hitting bad or partial data', 'Yes — a figure that cannot be computed now shows an em dash, never a fake zero',
     'The money test now pins the em dash, with the reason written next to it'],
    ['D6', 'The audit harness itself was lying. Its "founder-shaped" data described debts with field names this app does not use, so every debt read as $0 and the dumps we were auditing showed a screen with no debts on it.',
     'Nobody directly — but every mock-match audit run against it was checking the wrong picture', 'Yes — real field names, and the fixture now reproduces the approved mock\'s arithmetic exactly',
     'The dump now reconciles to $813,152 − $418,000 = $395,152'],
], [6, 76, 44, 46, 58], fills=lambda k, row: WARN)

# ── 5 Open founder calls (ANSWERED 2026-08-11) ───────────────────────────────
sheet(wb, 'Open founder calls', ['#', 'The question', 'What I did meanwhile', 'Why I did not just decide', 'Decision', 'Status'], [
    ['Q1', 'The cash row inside CASH: "Vanguard — sweep cash", or the account\'s real name with "cash in this account" beneath it?',
     'Kept the real account name', 'The mock\'s shorter name is friendlier; the built one is traceable to an actual account you can open.',
     'Keep the real account name', 'DONE — no code change; the mock was corrected to match'],
    ['Q2', 'Inside INVESTMENTS, order the categories biggest-first (the mock) or by liquidity (the rule approved in the pre-48 audit)?',
     'Kept liquidity order — the approved rule', 'Changing an approved ordering rule on the strength of one drawing is how rules quietly die.',
     'Biggest first', 'DONE — bar and list both sort by size; supersedes the liquidity rule, logged'],
    ['Q3', 'The "This month\'s cash flow" row and the "＋ Add or connect an account" button sit below the cushion. Neither is in the mock. Keep or drop?',
     'Kept both', 'They are approved features and the add button is this screen\'s only way to add an account.',
     'Keep add or connect. Delete cashflow.', 'DONE — the row and its code are gone; the button stays'],
    ['Q4', 'Should the By category / By institution buttons appear on the first-day screen?',
     'Left them off until there is an account', 'On an empty screen the second button changes nothing — a dead control on the first screen a new person sees.',
     'Yes, as it will give user view of what they can do in the app', 'DONE — both buttons show; By institution says what it will hold'],
    ['Q5', 'The handoff turns the whole hero into a solid green block. Keep your approved white hero card?',
     'Kept your approved hero', 'v1 in the mock folder shows the handoff version to compare.',
     'Keep the white hero card with green band', 'DONE — no change'],
    ['Q6', 'The handoff adds a composition bar under WHAT YOU OWE and a progress bar in the cushion. Neither is in the approved MoneyKeel mocks.',
     'Did not add either', 'They are new chart furniture, not corrections — worth a look before they go in.',
     'add both. Bar replaces donut.', 'DONE — donut removed; assets bar, debts bar and cushion bar built'],
], [6, 58, 36, 58, 40, 52], fills=lambda k, row: BAND)

# ── 6 Appearance audit ───────────────────────────────────────────────────────
sheet(wb, 'Appearance audit', ['Check (standing order step 5)', 'Method', 'Result'], [
    ['Band colours', 'Pulled the rendered colour off the element, not the stylesheet', 'PASS — section bands deep green #085041 with white caps; group bars light green with deep-green text'],
    ['Font sizes and weights', 'Walked every rendered node on the screen and compared each size to the design scale', 'PASS — zero off-scale sizes (11/13/15/17/20/24/30/38 only)'],
    ['THE shared right edge', 'Compared the rendered right inset of a section total, a group total, a class total and an account amount on the built screen',
     'PASS — all four resolve to the same edge; every row reserves the same 16-point arrow column, and rows that do not navigate hold it open'],
    ['Spacing, padding, indent', 'Rendered padding values', 'PASS — one row inset for the whole ledger; child rows indent exactly one step (never two)'],
    ['The donut', 'Rendered slice colours, order and labels', 'PASS — validated palette in the fixed class order, names and percentages beside the ring, small slices named on their own line, centre still reads Assets + total'],
    ['Order', 'Position of every section in the rendered string dump', 'PASS — buttons above the lists, hero → banner → plan → buttons → own → owe → cushion'],
    ['Nesting', 'Searched the rendered tree for a duplicated or wrongly parented heading', 'PASS — no class heading repeats its group name'],
    ['Data reality', 'Rendered with connected accounts that send no holdings', 'PASS — the CD/Treasuries account reads as Bonds & CDs; "Unclassified" appears nowhere on the screen'],
    ['NOT CHECKED — how it looks on the actual phone', 'A test renders the tree, it does not draw pixels on an iPhone',
     'Outstanding: the donut labels are placed by maths at a fixed 330-point width, so a long class name on the smallest screen is the thing to eye first. The way to check is a TestFlight build or a screenshot from the running app — not another test.'],
], [40, 56, 96])

wb.save(OUT)
print('wrote', OUT)
